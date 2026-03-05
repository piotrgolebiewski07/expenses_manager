import csv
from datetime import date, datetime, time
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from sqlalchemy import func
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session, joinedload

from app.core.exception import (
    CategoryNotFoundException,
    DatabaseException,
    ExpenseNotFoundException,
    InvalidMonthException,
    InvalidYearException,
    NoExpensesFoundException,
    UserAlreadyExistsException,
)
from app.core.security import hash_password
from app.models.models import Category, Expense, User
from app.schemas.schemas import ExpenseCreateDTO, ExpenseUpdateDTO, UserCreate


def get_all_expenses(db: Session,
                     current_user: User,
                     limit: int,
                     offset: int,
                     sort_by: str,
                     order: str,
                     min_price: int | None,
                     max_price: int | None,
                     start_date: date | None,
                     end_date: date | None,
                     category_id: int | None,
                     category_name: str | None
                     ):
    query = db.query(Expense).filter(Expense.user_id == current_user.id)

    # filter by price range
    if min_price is not None:
        query = query.filter(Expense.price >= min_price)
    if max_price is not None:
        query = query.filter(Expense.price <= max_price)

    if start_date is not None:
        start_dt = datetime.combine(start_date, time.min)
        query = query.filter(Expense.created_at >= start_dt)

    if end_date is not None:
        end_dt = datetime.combine(end_date, time.max)
        query = query.filter(Expense.created_at <= end_dt)

    if category_name is not None:
        query = query.join(Category).filter(Category.name == category_name)

    # filter by category
    if category_id is not None:
        query = query.filter(Expense.category_id == category_id)

    # dynamic sorting
    columns = {
        "id": Expense.id,
        "name": Expense.name,
        "price": Expense.price,
        "created_at": Expense.created_at
    }

    column = columns.get(sort_by, Expense.created_at)

    if order == "desc":
        query = query.order_by(column.desc(), Expense.id.desc())
    else:
        query = query.order_by(column.asc(), Expense.id.asc())

    total = query.order_by(None).count()

    items = (
        query
        .offset(offset)
        .limit(limit)
        .all()
    )

    return {
        "items": items,
        "total": total,
        "limit": limit,
        "offset": offset
    }


def get_expense_by_id(db: Session, expense_id: int, current_user: User):
    expense = db.query(Expense).filter(Expense.id == expense_id, Expense.user_id == current_user.id).first()
    if not expense:
        raise ExpenseNotFoundException()

    return expense


def create_expense(db: Session, dto: ExpenseCreateDTO, current_user: User):
    category = db.query(Category).filter(Category.id == dto.category_id).first()
    if not category:
        raise CategoryNotFoundException()

    expense = Expense(
        name=dto.name,
        price=dto.price,
        category=category,
        user_id=current_user.id
    )

    db.add(expense)
    db.commit()
    db.refresh(expense)

    return expense


def update_expense(db: Session, expense_id: int, dto: ExpenseUpdateDTO, current_user: User):
    expense = get_expense_by_id(db, expense_id, current_user)
    if dto.name is not None:
        expense.name = dto.name
    if dto.category_id is not None:
        category = db.query(Category).filter(Category.id == dto.category_id).first()
        if not category:
            raise CategoryNotFoundException()
        expense.category = category
    if dto.price is not None:
        expense.price = dto.price

    db.commit()
    db.refresh(expense)

    return expense


def delete_expense(db: Session, expense_id: int, current_user: User):
    expense = get_expense_by_id(db, expense_id, current_user)
    db.delete(expense)
    db.commit()

    return expense


def statistics(db: Session, year:int, month: int, current_user: User):
    if month < 1 or month > 12:
        raise InvalidMonthException()

    if year < 2000 or year > 2100:
        raise InvalidYearException()

    result = db.query(
        func.sum(Expense.price),
        func.avg(Expense.price),
        func.max(Expense.price),
        func.count(Expense.id)
    ).filter(
        func.strftime("%m", Expense.created_at) == f"{month:02}",
        func.strftime("%Y", Expense.created_at) == str(year),
        Expense.user_id == current_user.id
    ).one()

    total, average, max_expense, count = result

    total = total or 0
    average = round(average or 0, 2)
    max_expense = max_expense or 0
    count = count or 0

    category_stats = (
        db.query(Category.name, func.sum(Expense.price))
        .join(Expense, Expense.category_id == Category.id)
        .filter(
            func.strftime("%m", Expense.created_at) == f"{month:02}",
            func.strftime("%Y", Expense.created_at) == str(year),
            Expense.user_id == current_user.id
        )
        .group_by(Category.name)
        .all()
    )

    by_category = [
        {"category": name, "total": category_total or 0}
        for name, category_total in category_stats
    ]

    return {
        "total": total,
        "average": average,
        "max": max_expense,
        "count": count,
        "by_category": by_category
    }


def generate_visualization(db: Session, year: int, month: int, current_user: User):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np

    if month < 1 or month > 12:
        raise InvalidMonthException()

    if year < 2000 or year > 2100:
        raise InvalidYearException()

    expenses = (db.query(Expense)
                .options(joinedload(Expense.category))
                .filter(
            func.strftime("%m", Expense.created_at) == f"{month:02}",
                    func.strftime("%Y", Expense.created_at) == str(year),
                    Expense.user_id == current_user.id
                )
                .all()
                )

    if not expenses:
        raise NoExpensesFoundException()

    # Grupowanie
    categories = {}
    for expense in expenses:
        name = expense.category.name
        categories[name] = categories.get(name, 0) + expense.price

    # Przygotowanie danych do wykresu
    labels = list(categories.keys())
    values = list(categories.values())
    total_sum = sum(values)

    # Wyróżnienie największej kategorii
    max_index = values.index(max(values))
    explode = [0.08 if i == max_index else 0 for i in range(len(values))]

    # Kolorystyka
    cmap = plt.get_cmap("tab20")
    colors = cmap(np.linspace(0, 1, len(labels)))

    fig, ax = plt.subplots(figsize=(5, 5))

    ax.pie(
        values,
        labels=labels,
        autopct=lambda pct: f"{pct:.1f}%\n({int(pct / 100. * total_sum)})",
        startangle=90,
        explode=explode,
        colors=colors,
        wedgeprops={"width": 0.4, "edgecolor": "white", "linewidth": 1.2},
        textprops={"fontsize": 10}
    )

    # Tytuł
    ax.set_title(
        f"Expenses distribution\n{month:02}/{year}",
        fontsize=14,
        weight="bold"
    )

    # Równe proporcje koła
    ax.axis("equal")

    image_stream = io.BytesIO()
    fig.savefig(image_stream, format="png", bbox_inches="tight", dpi=150)
    plt.close(fig)
    image_stream.seek(0)

    return image_stream


def generate_report(
        db: Session,
        category: str | None,
        start_date: date | None,
        end_date: date | None,
        current_user: User
):
    query = (
        db.query(Expense)
        .options(joinedload(Expense.category))
        .filter(Expense.user_id == current_user.id)
    )

    if category:
        query = query.join(Category).filter(Category.name == category)

    if start_date:
        start_dt = datetime.combine(start_date, time.min)
        query = query.filter(Expense.created_at >= start_dt)

    if end_date:
        end_dt = datetime.combine(end_date, time.max)
        query = query.filter(Expense.created_at <= end_dt)

    expenses = query.order_by(Expense.created_at.asc()).all()

    if not expenses:
        raise NoExpensesFoundException()

    wb = Workbook()

    # =========================
    # SHEET 1 — DATA
    # =========================
    ws_data = wb.active
    ws_data.title = "Data"

    headers = ["ID", "Name", "Category", "Price", "Created At"]
    ws_data.append(headers)

    for col in range(1, len(headers) + 1):
        ws_data.cell(row=1, column=col).font = Font(bold=True)

    for exp in expenses:
        ws_data.append([
            exp.id,
            exp.name,
            exp.category.name if exp.category else None,
            exp.price,
            exp.created_at
        ])

    ws_data.column_dimensions["A"].width = 8
    ws_data.column_dimensions["B"].width = 22
    ws_data.column_dimensions["C"].width = 18
    ws_data.column_dimensions["D"].width = 14
    ws_data.column_dimensions["E"].width = 20

    for row in ws_data.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = "#,##0.00"

    # =========================
    # SHEET 2 — SUMMARY
    # =========================
    ws_summary = wb.create_sheet("Summary")

    # Nagłówek
    ws_summary["A1"] = "EXPENSE REPORT SUMMARY"
    ws_summary.merge_cells("A1:B1")
    ws_summary["A1"].font = Font(size=16, bold=True)
    ws_summary["A1"].alignment = Alignment(horizontal="center")
    ws_summary["A1"].fill = PatternFill(
        start_color="DDDDDD",
        end_color="DDDDDD",
        fill_type="solid"
    )

    # Okres raportu
    period_text = f"Period: {start_date or '---'} - {end_date or '---'}"
    ws_summary["A2"] = period_text
    ws_summary.merge_cells("A2:B2")
    ws_summary["A2"].alignment = Alignment(horizontal="center")

    # Podstawowe statystyki
    total = sum(exp.price for exp in expenses)
    count = len(expenses)
    average = total / count if count else 0
    max_value = max(exp.price for exp in expenses)

    summary_rows = [
        ("Total", total),
        ("Average", average),
        ("Max", max_value),
        ("Count", count),
    ]

    for idx, (label, value) in enumerate(summary_rows, start=4):
        ws_summary[f"A{idx}"] = label
        ws_summary[f"B{idx}"] = value
        ws_summary[f"A{idx}"].font = Font(bold=True)

    # Format waluty (Total, Average, Max)
    for row in ws_summary.iter_rows(min_row=4, max_row=6, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = "#,##0.00"

    # Ramki dla summary
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws_summary.iter_rows(min_row=4, max_row=7, min_col=1, max_col=2):
        for cell in row:
            cell.border = border

    # Sekcja kategorii
    ws_summary["A9"] = "By Category"
    ws_summary["A9"].font = Font(size=12, bold=True)

    category_totals = {}
    for exp in expenses:
        name = exp.category.name
        category_totals[name] = category_totals.get(name, 0) + exp.price

    # Sortowanie malejąco
    sorted_categories = sorted(
        category_totals.items(),
        key=lambda x: x[1],
        reverse=True
    )

    row_start = 11

    for idx, (cat, value) in enumerate(sorted_categories):
        ws_summary[f"A{row_start + idx}"] = cat
        ws_summary[f"B{row_start + idx}"] = value
        ws_summary[f"B{row_start + idx}"].number_format = "#,##0.00"

        # Wyróżnienie największej kategorii
        if idx == 0:
            ws_summary[f"A{row_start + idx}"].font = Font(bold=True)
            ws_summary[f"B{row_start + idx}"].font = Font(bold=True)

        ws_summary[f"A{row_start + idx}"].border = border
        ws_summary[f"B{row_start + idx}"].border = border

    ws_summary.column_dimensions["A"].width = 22
    ws_summary.column_dimensions["B"].width = 16

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def create_user(db: Session, user: UserCreate):
    existing_user = db.query(User).filter(User.email == user.email).first()

    if existing_user:
        raise UserAlreadyExistsException()

    hashed_pw = hash_password(user.password)

    db_user = User(
        email=user.email,
        hashed_password=hashed_pw
    )

    db.add(db_user)
    db.commit()
    db.refresh(db_user)

    return db_user


def get_user_by_email(db: Session, email: str):
    return db.query(User).filter(User.email == email).first()

